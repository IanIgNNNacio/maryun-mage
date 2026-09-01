"""Genera override_clasificacion.xlsx desde el archivo externo ABC/XYZ de Maryun.

Lee ``RESULTADO_ESTRATEGICO_ABC_XYZ_v3.xlsx`` (o la ruta indicada con --input),
hace match exacto de nombres de producto contra el catálogo del pipeline y escribe
``data/reference/override_clasificacion.xlsx`` para que Stage 5 (overrides) lo
aplique en la próxima corrida.

Toda variante (talla, color) que comparte el mismo ``nombre`` de producto en una
sucursal recibe la misma clasificación ABC/XYZ del archivo externo.

Uso rápido:
    cd C:\\Users\\user\\Desktop\\maryun-replenishment
    python src/app/classification/external_abc_xyz.py
    python src/app/classification/external_abc_xyz.py --input "ruta/al/RESULTADO_ESTRATEGICO.xlsx"
    python src/app/classification/external_abc_xyz.py --dry-run
"""
from __future__ import annotations

import argparse
import shutil
import sys
from pathlib import Path

import pandas as pd

# Ruta raíz del proyecto — asegura imports desde src/
_ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(_ROOT / "src"))

from app.normalize.canonical import SUCURSALES_CANONICAS

# ---------------------------------------------------------------------------
# Rutas por defecto
# ---------------------------------------------------------------------------
_DEFAULT_INPUT = (
    Path(r"C:\Users\user\OneDrive\Escritorio\DEMANDA 2024-2025")
    / "RESULTADO_ESTRATEGICO_ABC_XYZ_v3.xlsx"
)
_REFERENCE_DIR   = _ROOT / "data" / "reference"
_BACKUP_EXTERNO  = _REFERENCE_DIR / "abc_xyz_externo.xlsx"
_OVERRIDE_OUT    = _REFERENCE_DIR / "override_clasificacion.xlsx"
_OVERRIDE_SHEET  = "override_clasificacion"

# Motivo estampado en cada fila generada (permite distinguirlas de overrides manuales)
_MOTIVO    = "externo_abc_xyz_v3"
_RESPONSABLE = "sistema"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _load_products_from_audit() -> pd.DataFrame | None:
    """Lee el catálogo de productos del último run_audit.xlsx."""
    output_dir = _ROOT / "data" / "output"
    if not output_dir.exists():
        return None
    # Solo carpetas que parezcan run_id (empiezan con dígito) y tengan audit
    runs = sorted([
        p for p in output_dir.iterdir()
        if p.is_dir() and p.name[:1].isdigit() and (p / "run_audit.xlsx").exists()
    ])
    if not runs:
        return None
    audit = runs[-1] / "run_audit.xlsx"
    if not audit.exists():
        return None
    try:
        df = pd.read_excel(audit, sheet_name="products", dtype=str)
        print(f"  📋  Catálogo: {audit.relative_to(_ROOT)}  ({len(df):,} productos)")
        return df
    except Exception as exc:
        print(f"  ⚠️  No se pudo leer products de {audit}: {exc}")
        return None


def _load_products_from_mariadb() -> pd.DataFrame | None:
    """Fallback: trae sku_id + nombre directamente de MariaDB."""
    try:
        env_file = _ROOT / ".env"
        if env_file.exists():
            try:
                from dotenv import load_dotenv
                load_dotenv(env_file, override=False)
            except ImportError:
                import os
                for line in env_file.read_text(encoding="utf-8").splitlines():
                    line = line.strip()
                    if not line or line.startswith("#") or "=" not in line:
                        continue
                    k, v = line.split("=", 1)
                    os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))

        from app.connectors.mariadb import MariaDBConnector
        conn = MariaDBConnector()
        df = conn.run_query("SELECT sku_id, nombre FROM articulos")
        df["sku_id"] = df["sku_id"].astype(str)
        df["nombre"] = df["nombre"].astype(str)
        print(f"  📋  Catálogo desde MariaDB: {len(df):,} productos")
        return df
    except Exception as exc:
        print(f"  ⚠️  MariaDB fallback falló: {exc}")
        return None


def _extract_nombre_base(df: pd.DataFrame) -> pd.DataFrame:
    """Extrae nombre_base de SKU 3.0 quitando el sufijo SUCURSAL.

    La columna ``SKU 3.0`` contiene ``nombre_producto + SUCURSAL`` concatenados
    sin separador. Se elimina exactamente ``len(SUCURSAL)`` caracteres del final.
    """
    df = df.copy()
    df["SKU 3.0"] = df["SKU 3.0"].astype(str).str.strip()
    df["SUCURSAL"] = df["SUCURSAL"].astype(str).str.strip()

    def _extract(row: pd.Series) -> str | None:
        sku3 = row["SKU 3.0"]
        suc  = row["SUCURSAL"]
        if sku3.endswith(suc):
            base = sku3[: -len(suc)].strip()
            return base if base else None
        return None

    df["nombre_base"] = df.apply(_extract, axis=1)
    n_fail = df["nombre_base"].isna().sum()
    if n_fail:
        print(f"  ⚠️  {n_fail} filas con extracción fallida (SKU 3.0 no termina en SUCURSAL)")
    return df


def _load_external(path: Path) -> pd.DataFrame:
    """Lee y valida el archivo externo ABC/XYZ."""
    print(f"  📂  Leyendo archivo externo: {path.name}")
    df = pd.read_excel(path, sheet_name="Analisis", dtype=str)
    required = {"SKU 3.0", "SUCURSAL", "ABC", "XYZ", "MATRIZ_FINAL"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Columnas faltantes en el archivo externo: {missing}")

    # Limpieza básica
    for col in ("ABC", "XYZ", "MATRIZ_FINAL"):
        df[col] = df[col].str.strip().str.upper()

    # Validar valores
    bad_abc = ~df["ABC"].isin(["A", "B", "C"])
    bad_xyz = ~df["XYZ"].isin(["X", "Y", "Z"])
    if bad_abc.any():
        print(f"  ⚠️  {bad_abc.sum()} filas con ABC inválido — se descartan")
        df = df[~bad_abc]
    if bad_xyz.any():
        print(f"  ⚠️  {bad_xyz.sum()} filas con XYZ inválido — se descartan")
        df = df[~bad_xyz]

    print(f"  ✅  Archivo externo: {len(df):,} filas válidas")
    return df


def _build_overrides(df_ext: pd.DataFrame, df_prod: pd.DataFrame) -> pd.DataFrame:
    """Cruza el archivo externo con el catálogo y genera las filas de override.

    Cada nombre de producto en el archivo externo puede corresponder a varios
    sku_ids en el pipeline (variantes de talla, color, etc.). Todos reciben la
    misma clasificación ABC/XYZ.
    """
    # ---- normalización para el join ----
    df_ext = df_ext.copy()
    df_ext["_nombre_upper"] = df_ext["nombre_base"].str.strip().str.upper()
    df_ext["_ubicacion_upper"] = df_ext["SUCURSAL"].str.strip().str.upper()

    df_prod = df_prod.copy()
    df_prod["_nombre_upper"] = df_prod["nombre"].astype(str).str.strip().str.upper()

    # ---- eliminar duplicados en externo (mismo nombre+ubicacion → mantener 1 fila) ----
    # En caso de conflicto, mantenemos el que tenga mejor ABC (A > B > C).
    _rank = {"A": 0, "B": 1, "C": 2}
    df_ext["_abc_rank"] = df_ext["ABC"].map(_rank).fillna(9)
    df_ext = df_ext.sort_values("_abc_rank")
    df_ext = df_ext.drop_duplicates(subset=["_nombre_upper", "_ubicacion_upper"], keep="first")

    # ---- join exacto por nombre ----
    merged = df_prod.merge(
        df_ext[["_nombre_upper", "_ubicacion_upper", "ABC", "XYZ", "MATRIZ_FINAL"]],
        on="_nombre_upper",
        how="inner",
    )

    # Si products no tiene 'ubicacion', no podemos filtrar por sucursal aquí —
    # pero cross join controlado: un SKU puede estar en múltiples sucursales.
    # El campo _ubicacion_upper del externo es la sucursal destino.
    # Renombramos para que el loader lo entienda.
    result = merged[["sku_id", "_ubicacion_upper", "ABC", "XYZ"]].copy()
    result = result.rename(columns={
        "_ubicacion_upper": "ubicacion",
        "ABC": "abc_override",
        "XYZ": "xyz_override",
    })
    result["motivo"]      = _MOTIVO
    result["responsable"] = _RESPONSABLE

    # Deduplicar (mismo sku_id + ubicacion — puede pasar si hay variantes duplicadas)
    result = result.drop_duplicates(subset=["sku_id", "ubicacion"], keep="first")

    return result[["sku_id", "ubicacion", "abc_override", "xyz_override", "motivo", "responsable"]]


def _load_existing_manual_overrides(path: Path) -> pd.DataFrame:
    """Lee overrides manuales existentes (los que NO tienen motivo externo)."""
    if not path.exists():
        return pd.DataFrame(columns=["sku_id", "ubicacion", "abc_override", "xyz_override", "motivo", "responsable"])
    try:
        df = pd.read_excel(path, sheet_name=_OVERRIDE_SHEET, dtype=str)
        # Filtrar solo las filas que NO fueron generadas por este script
        if "motivo" in df.columns:
            manual = df[df["motivo"].fillna("") != _MOTIVO].copy()
        else:
            manual = df.copy()
        if not manual.empty:
            print(f"  🔒  Se conservan {len(manual)} overrides manuales existentes")
        return manual
    except Exception:
        return pd.DataFrame(columns=["sku_id", "ubicacion", "abc_override", "xyz_override", "motivo", "responsable"])


def _write_override_file(df: pd.DataFrame, path: Path) -> None:
    """Escribe el DataFrame en el sheet requerido por el pipeline."""
    path.parent.mkdir(parents=True, exist_ok=True)
    # Si el archivo ya existe y tiene otros sheets, los preservamos
    if path.exists():
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path)
            if _OVERRIDE_SHEET in wb.sheetnames:
                del wb[_OVERRIDE_SHEET]
            with pd.ExcelWriter(path, engine="openpyxl", mode="a") as writer:
                writer.book = wb
                df.to_excel(writer, sheet_name=_OVERRIDE_SHEET, index=False)
            return
        except Exception:
            pass  # Si falla, sobreescribimos el archivo completo

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=_OVERRIDE_SHEET, index=False)


# ---------------------------------------------------------------------------
# Función principal pública (importable por el pipeline)
# ---------------------------------------------------------------------------

def generate_classification_overrides(
    input_path: Path | None = None,
    dry_run: bool = False,
    verbose: bool = True,
) -> pd.DataFrame:
    """Genera overrides de clasificación desde el archivo externo.

    Args:
        input_path: Ruta al archivo RESULTADO_ESTRATEGICO_ABC_XYZ_v3.xlsx.
                    Si es None, usa la ruta por defecto (_DEFAULT_INPUT).
        dry_run:    Si True, solo muestra el resultado sin escribir archivos.
        verbose:    Si True, imprime mensajes en consola.

    Returns:
        DataFrame con las filas de override generadas.
    """
    def _log(*args):
        if verbose:
            print(*args)

    path = Path(input_path) if input_path else _DEFAULT_INPUT
    if not path.exists():
        _log(f"  ❌  Archivo externo no encontrado: {path}")
        _log("       Pasá la ruta con --input <ruta>")
        return pd.DataFrame()

    # 1. Leer y parsear archivo externo
    df_ext = _load_external(path)
    df_ext = _extract_nombre_base(df_ext)
    df_ext = df_ext[df_ext["nombre_base"].notna()]

    # 2. Cargar catálogo de productos
    df_prod = _load_products_from_audit()
    if df_prod is None:
        _log("  🔌  Intentando MariaDB como fuente de productos…")
        df_prod = _load_products_from_mariadb()
    if df_prod is None or df_prod.empty:
        _log("  ❌  No se pudo obtener el catálogo de productos.")
        return pd.DataFrame()

    # 3. Construir overrides
    df_ov = _build_overrides(df_ext, df_prod)

    # 4. Estadísticas de cobertura
    ext_nombres_set  = set(df_ext["nombre_base"].str.strip().str.upper().dropna())
    prod_nombres_set = set(df_prod["nombre"].str.strip().str.upper().dropna())
    matched_set      = ext_nombres_set & prod_nombres_set
    unmatched_set    = ext_nombres_set - prod_nombres_set

    _log(f"\n  ── Cobertura ────────────────────────────────────────")
    _log(f"  Nombres en archivo externo : {len(ext_nombres_set):>6,}")
    _log(f"  Nombres con match exacto   : {len(matched_set):>6,} ({len(matched_set)/max(len(ext_nombres_set),1)*100:.1f}%)")
    _log(f"  Sin match en pipeline      : {len(unmatched_set):>6,} (productos sin historial en periodo actual)")
    _log(f"  Overrides generados (filas): {len(df_ov):>6,}  (variantes/tallas incluidas)")

    # Distribución por clase
    clase_dist = (df_ov["abc_override"] + df_ov["xyz_override"]).value_counts().head(9)
    _log("\n  ── Top clases ABC/XYZ ───────────────────────────────")
    for clase, n in clase_dist.items():
        _log(f"  {clase:<4}  {n:>6,} pares sku/sucursal")

    if dry_run:
        _log("\n  ℹ️  Modo dry-run: no se escribieron archivos.")
        return df_ov

    # 5. Preservar overrides manuales y fusionar
    df_manual = _load_existing_manual_overrides(_OVERRIDE_OUT)
    if not df_manual.empty:
        # Asegurar columnas compatibles
        for col in ("abc_override", "xyz_override", "motivo", "responsable"):
            if col not in df_manual.columns:
                df_manual[col] = pd.NA
        df_final = pd.concat([df_ov, df_manual], ignore_index=True)
        # En caso de conflicto (mismo sku+ubicacion), el manual gana
        df_final = df_final.drop_duplicates(subset=["sku_id", "ubicacion"], keep="last")
    else:
        df_final = df_ov

    # 6. Copia de respaldo del archivo externo en data/reference/
    if not dry_run:
        try:
            _REFERENCE_DIR.mkdir(parents=True, exist_ok=True)
            shutil.copy2(path, _BACKUP_EXTERNO)
            _log(f"\n  💾  Respaldo guardado: {_BACKUP_EXTERNO.relative_to(_ROOT)}")
        except Exception as exc:
            _log(f"  ⚠️  No se pudo copiar respaldo: {exc}")

    # 7. Escribir override file
    _write_override_file(df_final, _OVERRIDE_OUT)
    _log(f"  ✅  Override generado: {_OVERRIDE_OUT.relative_to(_ROOT)}")
    _log(f"      {len(df_final):,} filas  ({len(df_ov):,} externas + {len(df_manual):,} manuales)")
    _log(f"\n  ▶️   Ejecutá el pipeline para aplicar las clasificaciones.")

    return df_final


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Genera override_clasificacion.xlsx desde el archivo externo ABC/XYZ."
    )
    parser.add_argument(
        "--input", "-i",
        type=Path,
        default=None,
        help="Ruta al archivo RESULTADO_ESTRATEGICO_ABC_XYZ_v3.xlsx",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Previsualiza sin escribir archivos",
    )
    return parser.parse_args()


if __name__ == "__main__":
    import io, os
    # Forzar UTF-8 en la salida de consola (Windows cp1252 no soporta todos los caracteres)
    if hasattr(sys.stdout, "reconfigure"):
        try:
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass

    args = _parse_args()
    print()
    print("  == Generador de Overrides ABC/XYZ Externo ==============")
    result = generate_classification_overrides(
        input_path=args.input,
        dry_run=args.dry_run,
    )
    if result.empty and not args.dry_run:
        sys.exit(1)
