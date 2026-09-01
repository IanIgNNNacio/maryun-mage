"""Validacion estricta de los archivos fuente OFICIALES.

Reglas de negocio (obligatorias):
  - Pronostico: SOLO desde PRONOSTICOS_DEMANDA_v10.xlsx (hoja 'Pronosticos').
  - Clasificacion: SOLO desde RESULTADO_ESTRATEGICO_ABC_XYZ_v3.xlsx (hoja 'Analisis').

Si el archivo no existe, no es accesible, no tiene la hoja esperada o le faltan
columnas, el proceso DEBE FALLAR de forma explicita (SourceFileError). NUNCA debe
continuar con copias locales, versiones antiguas, temporales o rutas alternativas.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd

from app.utils.logging import get_logger

# ---------------------------------------------------------------------------
# Rutas OFICIALES — unica fuente de verdad. No usar copias ni versiones previas.
# ---------------------------------------------------------------------------
_BASE_DEMANDA = Path(r"C:\Users\user\OneDrive\Escritorio\DEMANDA 2024-2025")
OFFICIAL_FORECAST = _BASE_DEMANDA / "PRONOSTICOS_DEMANDA_v10.xlsx"
OFFICIAL_ABC_XYZ = _BASE_DEMANDA / "RESULTADO_ESTRATEGICO_ABC_XYZ_v3.xlsx"

# Hojas y columnas esperadas — si cambian, la validacion falla (encabezados alterados).
FORECAST_SHEET = "Pronosticos"
FORECAST_REQUIRED_COLS = {"ID_KEY", "Fecha", "Pronostico"}

CLASIF_SHEET = "Analisis"
CLASIF_REQUIRED_COLS = {"SKU 3.0", "ABC", "XYZ", "MATRIZ_FINAL"}
# SUCURSAL fue eliminado del archivo v3 en jun-2026; se auto-detecta desde SKU 3.0.


class SourceFileError(RuntimeError):
    """Falla explicita de validacion de un archivo fuente oficial."""


def _count_rows(path: Path, sheet: str) -> int | None:
    """Cuenta filas de datos (sin encabezado) de forma eficiente (read_only)."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        try:
            ws = wb[sheet]
            n = ws.max_row
            return (n - 1) if (n is not None and n > 0) else None
        finally:
            wb.close()
    except Exception:
        return None


def _validate(
    path: Path,
    official: Path,
    expected_sheet: str,
    required_cols: set[str],
    label: str,
) -> dict:
    """Valida un archivo fuente oficial. Lanza SourceFileError ante cualquier fallo.

    Returns:
        dict con metadata: ruta, mtime, hoja, columnas, filas_excel.
    """
    log = get_logger("validation.source_files")

    # 1. Coincidencia con la ruta oficial (no rutas alternativas/antiguas)
    try:
        same_as_official = path.resolve() == official.resolve()
    except Exception:
        same_as_official = str(path) == str(official)
    if not same_as_official:
        raise SourceFileError(
            f"[{label}] La ruta configurada NO es la oficial.\n"
            f"  configurada: {path}\n"
            f"  oficial    : {official}\n"
            f"  -> Corrige MARYUN_EXTERNAL__* en el .env. No se permiten copias/versiones."
        )

    # 2. Existencia
    if not path.exists():
        raise SourceFileError(
            f"[{label}] Archivo OFICIAL no encontrado: {path}\n"
            f"  El proceso no puede continuar sin la fuente oficial."
        )

    # 3. Accesibilidad (lectura)
    try:
        with open(path, "rb") as fh:
            fh.read(64)
    except Exception as exc:
        raise SourceFileError(f"[{label}] Archivo no accesible: {path} ({exc})") from exc

    # 4. Hoja correcta
    try:
        xl = pd.ExcelFile(path)
        sheets = xl.sheet_names
    except Exception as exc:
        raise SourceFileError(f"[{label}] No se pudo abrir como Excel: {path} ({exc})") from exc

    if expected_sheet not in sheets:
        raise SourceFileError(
            f"[{label}] Falta la hoja esperada '{expected_sheet}'. Hojas presentes: {sheets}"
        )

    # 5. Columnas/encabezados correctos (sin cambios)
    head = pd.read_excel(path, sheet_name=expected_sheet, nrows=0)
    cols = set(head.columns)
    missing = required_cols - cols
    if missing:
        raise SourceFileError(
            f"[{label}] Encabezados alterados — faltan columnas {sorted(missing)} "
            f"en hoja '{expected_sheet}'. Columnas presentes: {sorted(cols)[:20]}..."
        )

    # 6. Fecha de modificacion + conteo de filas
    st = path.stat()
    mtime_iso = pd.Timestamp(st.st_mtime, unit="s").isoformat()
    n_rows = _count_rows(path, expected_sheet)

    meta = {
        "label": label,
        "ruta": str(path),
        "es_oficial": True,
        "hoja": expected_sheet,
        "columnas_ok": True,
        "mtime": mtime_iso,
        "tamano_bytes": st.st_size,
        "filas_excel": n_rows,
    }
    log.info(
        "validation.source_ok",
        archivo=label,
        ruta=str(path),
        hoja=expected_sheet,
        mtime=mtime_iso,
        filas_excel=n_rows,
    )
    return meta


def validate_forecast_source(path: Path) -> dict:
    """Valida el Excel oficial de pronostico de demanda."""
    return _validate(path, OFFICIAL_FORECAST, FORECAST_SHEET,
                     FORECAST_REQUIRED_COLS, "PRONOSTICO_v10")


def validate_classification_source(path: Path) -> dict:
    """Valida el Excel oficial de clasificacion ABC/XYZ v3."""
    return _validate(path, OFFICIAL_ABC_XYZ, CLASIF_SHEET,
                     CLASIF_REQUIRED_COLS, "CLASIFICACION_v3")
